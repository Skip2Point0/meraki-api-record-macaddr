import requests
import json
from openpyxl import load_workbook

# ################################################# PARAMETERS BELOW ###################################################
# Please fill in parameters below.
# You can add devices to one or multiple networks.
# ap_abr is a variable that is used to track device names. Typical syntax for NN1-AP1-25 is NN1-
# Make sure that every parameter is added in correct order: Network Name1 > Sheet Name1 > NN1-
# ######################################################################################################################
meraki_api = 'APIKEY'
organization_id = 'Organization Name'
network_ids = ['Network Name1', 'Network Name2']
spread = 'spreadsheet.xlsx'
tabs = ['Sheet Name1', 'Sheet Name2']
ap_abr = ['NN1-', 'NN2-']
# ##################################################### PARAMETERS #####################################################
# If using default spreadsheet, no changes below are necessary.
# Custom spreadsheets require column changes below.
# ######################################################################################################################
ap_name_column = 'A'
serials_column = 'D'
mac_column = 'E'
# ################################################# PARAMETERS ABOVE ###################################################
net_dictionary = {}
shard_url = ()
headers = {
    'X-Cisco-Meraki-API-Key': meraki_api,
    'Content-Type': 'application/json'
}
print(spread)
wb = load_workbook(spread)


def pull_organization_id(head):
    global shard_url
    url = "https://api.meraki.com/api/v0/organizations"
    payload = {}
    response = requests.request("GET", url, headers=head, data=payload)
    response = response.content
    response = json.loads(response)
    for dicti in response:
        name = dicti["name"]
        if name == organization_id:
            org_id = dicti["id"]
            shard_url = dicti["url"]
            urlLenght = shard_url.find('com') + 3
            shard_url = shard_url[:urlLenght]
            print("#################################################")
            print(name + "\n" + "Organization ID: " + org_id)
            print("Organization Shard URL: " + shard_url)
            print("#################################################")
            return org_id
        else:
            continue


def pull_organization_networks(head):
    global net_dictionary
    global organization_id

    organization_id = pull_organization_id(head)
    url = shard_url + "/api/v0/organizations/" + organization_id + "/networks"
    payload = {}
    response = requests.request("GET", url, headers=head, data=payload)
    response = response.content
    json_response = json.loads(response)
    for networks in json_response:
        name = networks['name']
        n_id = networks['id']
        net_dictionary[name] = n_id
    print(net_dictionary)
    return net_dictionary


def pull_destination_networks():
    global network_ids

    dest_network_ids = []
    for n in network_ids:
        for i in net_dictionary:
            if n == i:
                print("Destination Network: " + n)
                dest_network_ids.append(net_dictionary[n])
                break
            else:
                continue
    print(dest_network_ids)
    return dest_network_ids


pull_organization_networks(headers)
networks_dest = pull_destination_networks()


def meraki_ap_parse_mac(workbook, networks, tbs, abr, name, serial, mac, head):
    ap_index = []
    print("########################################")
    print("PARSING FOR MAC AND SAVING...")
    incr = 0
    for tab in tbs:
        sheet = workbook[tab]
        # unsorted_ap = []
        row_index = []
        print("########################################")
        print("NETWORK:  " + tab)
        print("########################################")

        incr2 = 1
        for row in range(sheet.max_row):
            a = sheet[name + str(incr2)]
            unsorted_ap = a.value
            incr2 = incr2 + 1

            if abr[incr] in str(unsorted_ap):
                n_ap_index = row + 1
                ap_index.append(unsorted_ap)
                row_index.append(row)
                serials = sheet[serial + str(n_ap_index)].value

                url = shard_url + "/api/v0/networks/" + networks[incr] + "/devices/" + serials

                payload = {
                    "serial": serials,
                }
                payload = json.dumps(payload)
                response = requests.request("GET", url, headers=head, data=payload)
                f = response.content
                json_data = json.loads(f)
                print(json_data)
                macs = json_data["mac"]
                names = json_data["name"]
                print(names + " : " + macs)
                sheet[mac + str(n_ap_index)] = macs
                wb.save(filename=spread)

        incr = incr + 1


meraki_ap_parse_mac(wb, networks_dest, tabs, ap_abr, ap_name_column, serials_column, mac_column, headers)
